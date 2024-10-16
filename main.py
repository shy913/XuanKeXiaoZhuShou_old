from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import requests
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

version = "4.12"

service = Service(service_args=['--log-level=OFF'], executable_path=r'chromedriver.exe')
options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors')

driver = webdriver.Chrome(service=service, options=options)
# driver = webdriver.Chrome(service=service)
# driver = webdriver.Chrome()
print("Made by shy")

driver.get("http://xk.autoisp.shu.edu.cn/StudentQuery/QueryCourseList")
driver.minimize_window()
term_num = ""


def read_file(url):
    path = url
    result = {}
    try:
        with open(path, "r") as f:
            student_info = f.read()
        # print(student_info.split("\n"))
        for item in student_info.split("\n"):
            result[item.split(":")[0]] = item.split(":")[1]
        return result
    except():
        return -1


def border(font):
    if font == 2:
        print("=================================")
    elif font == 1:
        print("—————————————————————————————————")


def get_remain(CID, TeachNo):
    try:
        response = requests.post("http://xk.autoisp.shu.edu.cn/StudentQuery/QueryCourseList",
                                 data={"CourseType": "B", "PageIndex": "1", "PageSize": "30", "TeachNo": TeachNo,
                                       "CID": CID},
                                 headers={"Cookie": coo(), "User-Agent": user_agent}, timeout=5).text
        soup = BeautifulSoup(response, "html.parser")
        class_name = soup.findAll("td")
        remain = int(class_name[8].string) - int(class_name[9].string)
        with open(r"log.txt", "a") as f:
            f.write(f'{str(time.strftime("%m月%d日%H:%M:%S"))}\n'
                    f'-查询:cid: {CID} tid:{TeachNo}\n'
                    f'-剩余:{remain}\n\n')
    except:
        print('get_remain请求超时！')
        with open(r"log.txt", "a") as f:
            f.write(f'{str(time.strftime("%m月%d日%H:%M:%S"))}\n'
                    f'-查询:cid: {CID} tid:{TeachNo}\n'
                    f'-请求失败\n\n')
        return -100
    return remain


def xk(CID, TID, try_full):
    border(2)
    print("正在查询：\n课程号：" + CID + "\n教师号：" + TID)
    remain = get_remain(CID, TID)
    print("剩余人数:" + str(remain))
    if (int(remain) > 0) or try_full == 'True':
        print("开始选课！")
        url = "http://xk.autoisp.shu.edu.cn/CourseSelectionStudent/CourseSelectionSave"
        headers = {
            "Accept": "*/*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Cookie": coo()
        }
        data = {"cids": CID, "tnos": TID}
        # credentials = "include"
        try:
            response = requests.post(url, headers=headers, data=data, cookies=None, allow_redirects=True, verify=True,
                                     timeout=5).text
        except:
            response = "寄"
            print("xk请求超时")
            with open(r"log.txt", "a") as f:
                f.write(f'{str(time.strftime("%m月%d日%H:%M:%S"))}\n'
                        f'-选课:cid: {CID} tid:{TID}\n'
                        f'-xk请求失败\n\n')
        if "成功" in response:
            print("success")
            with open(r"log.txt", "a") as f:
                f.write(f'{str(time.strftime("%m月%d日%H:%M:%S"))}\n'
                        f'-选课:cid: {CID} tid:{TID}\n'
                        f'-选课成功\n\n')
            return "success"
        else:
            print("failed")
            with open(r"log.txt", "a") as f:
                f.write(f'{str(time.strftime("%m月%d日%H:%M:%S"))}\n'
                        f'-选课:cid: {CID} tid:{TID}\n'
                        f'-选课失败\n\n')
            return -1
    else:
        print("人数已满！")
        with open(r"log.txt", "a") as f:
            f.write(f'{str(time.strftime("%m月%d日%H:%M:%S"))}\n'
                    f'-选课:cid: {CID} tid:{TID}\n'
                    f'-人数已满\n\n')
        return 0


def get_cookie():
    # print("正在验证账号...请稍后！\n正在输入学号")
    form_student_id = driver.find_element(By.ID, "username")
    form_student_id.send_keys(student_id)
    # print("正在输入密码")
    form_password = driver.find_element(By.ID, "password")
    form_password.send_keys(password)
    # print("正在尝试登录")
    driver.find_element(By.ID, "submit-button").click()
    time.sleep(sleep_time)

    # print("登录成功")
    term_num = 0
    terms = driver.find_elements(By.NAME, "rowterm")
    terms[term_num].click()
    driver.find_element(By.XPATH, "/html/body/form/div/button").click()
    # print("正在获取cookie")
    coo = "ASP.NET_SessionId=" + driver.get_cookies()[0]["value"]
    print("写入cookie:" + coo)
    with open(r"cookie.txt", "w") as f:
        f.write(coo)
    # print("写入配置成功")
    driver.find_element(By.XPATH, "/html/body/div[2]/header/nav/div/ul/li[5]/a").click()
    # border(1)


def print_class_list():
    wb = load_workbook("Info.xlsx")
    ws = wb.active
    print("编号\t课程号    \t教师号")
    for i in range(1, 10):
        classid = str(ws.cell(1, i).value)
        teacherid = str(ws.cell(2, i).value)
        if classid == "None" or teacherid == "None":
            break
        print(str(i) + "\t" + classid + "\t" + teacherid)


def coo():
    path_r = r"cookie.txt"
    with open(path_r, "r") as f:  # 打开文件
        coo = str(f.read())  # 读取文件
    return coo


student_info = read_file("Info.txt")
student_id = student_info["student_id"]
password = student_info["password"]
user_agent = student_info["user_agent"]
refresh_rate = float(student_info["refresh_rate"])
sleep_time = float(student_info["sleep_time"])
try_full = str(student_info["try_full"])
# print("获取本地配置成功")

### start
# get_cookie()
while 1:
    border(1)
    print("选课小助手 VERSION " + version)
    print("账号：" + student_id)
    print("密码：" + password[0] + "*******")
    # border(1)
    # print_class_list()
    border(1)
    print("[1]  " + "\t" + "查看课程\n"
    "[2]  " + "\t" + "更改选课\n"
    "[3]  " + "\t" + "查看配置\n"
    "[4]  " + "\t" + "手动更新cookie\n"
    "[5]  " + "\t" + "查看log\n"
    "[0]  " + "\t" + "退出\n"
    "[enter]" + "\t" + "开始选课")
    border(1)
    jumpto = input("请输入选项:")

    if jumpto == "5":
        with open(r"log.txt", "r") as f:
            clear = input(f.read() + "\n是否清空？\n[y]\t:清空\n[enter]\t:不清空\n请输入选项：")
        if clear == "y":
            with open(r"log.txt", "w") as f:
                f.write("")
    elif jumpto == "4":
        get_cookie()

    elif jumpto == "2":
        wb = load_workbook("Info.xlsx")
        ws = wb.active
        border(1)
        print_class_list()
        col = int(input("请输入编号："))
        print("原课程号为：" + str(ws.cell(1, col).value))
        ws.cell(1, col).value = input("请输入课程号:")
        print("原教师号为：" + str(ws.cell(2, col).value))
        ws.cell(2, col).value = input("请输入教师号:")
        border(1)
        wb.save(r"Info.xlsx")

    elif jumpto == "3":
        border(1)
        for i in student_info:
            print(str(i) + " \t:" + student_info[i])
        border(1)

    elif jumpto == "1":
        # dl
        border(1)
        form_student_id = driver.find_element(By.ID, "username")
        form_student_id.send_keys(student_id)
        # print("正在输入密码")
        form_password = driver.find_element(By.ID, "password")
        form_password.send_keys(password)
        # print("正在尝试登录")
        driver.find_element(By.ID, "submit-button").click()
        time.sleep(sleep_time)
        term_num = 0
        terms = driver.find_elements(By.NAME, "rowterm")
        terms[term_num].click()
        driver.find_element(By.XPATH, "/html/body/form/div/button").click()
        # \dl
        wb = load_workbook("Info.xlsx")
        ws = wb.active
        print("课程名称  \t课程号    \t教师号\t剩余人数\t上课时间")
        for i in range(1, 11):
            if str(ws.cell(1, i).value) == "None" or str(ws.cell(2, i).value) == "None":
                break
            else:
                # get response
                response = requests.post("http://xk.autoisp.shu.edu.cn/StudentQuery/QueryCourseList",
                                         data={"CourseType": "B", "PageIndex": "1", "PageSize": "10",
                                               "TeachNo": str(ws.cell(2, i).value),
                                               "CID": str(ws.cell(1, i).value)},
                                         headers={"Cookie": coo(), "User-Agent": user_agent})
                soup = BeautifulSoup(response.text, "html.parser")
                # print(response.text)
                class_name = soup.findAll("td")
                print(class_name[1].string + "\t"
                      + str(ws.cell(1, i).value) + "\t"
                      + str(ws.cell(2, i).value) + "\t"
                      + str(int(class_name[8].string) - int(class_name[9].string)) + "    \t"
                      + class_name[6].string)
        time.sleep(sleep_time)
        # log out

        driver.find_element(By.XPATH, "/html/body/div[2]/header/nav/div/ul/li[5]/a").click()
        # \log out
    elif jumpto == "0":
        exit(0)

    elif jumpto == "":
        print("开始登录！\n正在输入学号")
        form_student_id = driver.find_element(By.ID, "username")
        form_student_id.send_keys(student_id)
        print("正在输入密码")
        form_password = driver.find_element(By.ID, "password")
        form_password.send_keys(password)
        print("正在登录")

        border(1)
        driver.find_element(By.ID, "submit-button").click()
        ## 获取学期列表.获取cookie

        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6", "Connection": "keep-alive",
            "Cookie": coo()}
        result = requests.get("http://xk.autoisp.shu.edu.cn/Home/TermIndex", headers=headers)
        soup = BeautifulSoup(result.text, "html.parser")
        terms = soup.findAll("tr", attrs={"name": "rowterm"})
        i = 1
        for term in terms:
            print("[" + str(i) + "]:")
            print(str(term.td.string.replace(" ", "").replace("\n", "")))
            i += 1
        border(1)
        ##进入

        if term_num == "":
            if i == 2:
                term_num = 0
                print("已自动选择学期")
            else:
                term_num = input("输入学期：")
                term_num = int(term_num) - 1

            if term_num == "":
                term_num = 0
        else:
            print("已选择学期" + str(term_num + 1))

        time.sleep(sleep_time)
        terms = driver.find_elements(By.NAME, "rowterm")
        terms[term_num].click()
        driver.find_element(By.XPATH, "/html/body/form/div/button").click()
        time.sleep(sleep_time)
        try:
            driver.find_element(By.XPATH, "/html/body/div[2]/aside/section/ul/li[2]/ul/li[1]").click()
            time.sleep(sleep_time)
        except:
            print("出现了点问题...应该是屏幕比例造成的")
        if "学生禁止选课" in driver.page_source or "选课时间未到" in driver.page_source:
            # border(1)
            print("坏咯：不在选课时间!")
            border(1)
            time.sleep(3)
            driver.find_element(By.XPATH, "/html/body/div[2]/header/nav/div/ul/li[5]/a").click()
        else:
            loopTime = 1
            # 开始选课了！

            # 所有flg=0
            wb = load_workbook("Info.xlsx")
            ws = wb.active
            for i in range(1, 20):
                ws.cell(3, i).value = "0"
            wb.save(r"Info.xlsx")
            while 1:

                wb = load_workbook("Info.xlsx")
                ws = wb.active

                for i in range(1, 11):
                    if str(ws.cell(1, i).value) == "None" or str(ws.cell(2, i).value) == "None":
                        break
                    wb = load_workbook("Info.xlsx")
                    ws = wb.active
                    # print(f'{i}:{ws.cell(3, i).value}')
                    if ws.cell(3, i).value == "0":
                        flg = xk(str(ws.cell(1, i).value), str(ws.cell(2, i).value), try_full)
                        if flg == "success":
                            # 修改flg = 1
                            wb = load_workbook("Info.xlsx")
                            ws = wb.active
                            ws.cell(3, i).value = "1"
                            wb.save(r"Info.xlsx")
                            print(f"嘿嘿嘿哈!\n")

                    else:
                        print("已跳过")
                        with open(r"log.txt", "a") as f:
                            f.write(f'{str(time.strftime("%m月%d日%H:%M:%S"))}\n'
                                    f'课程号{str(ws.cell(1, i).value)}, 教师号{str(ws.cell(2, i).value)}\n'
                                    f'已跳过\n\n')
                    border(2)
                time.sleep(sleep_time)
                print("第" + str(loopTime) + "次循环已结束！")
                print(time.strftime("%m月%d日%H:%M:%S"))
                with open(r"log.txt", "a") as f:
                    f.write(f"{str(time.strftime("%m月%d日%H:%M:%S"))}\n"
                            f"-第{str(loopTime)}次查询结束\n"
                            f"===========================\n\n")
                time.sleep(refresh_rate - sleep_time)
                loopTime += 1
